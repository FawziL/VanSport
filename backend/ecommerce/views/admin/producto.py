# pip install openpyxl

from rest_framework import viewsets, permissions, status
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from django.conf import settings
import os
import json
from ecommerce.models import Producto
from ecommerce.serializers import ProductoSerializer
from rest_framework.decorators import action
from rest_framework.permissions import IsAdminUser
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from decimal import Decimal
from datetime import datetime

class ProductoViewSetAdmin(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    permission_classes = [permissions.IsAdminUser]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def _parse_urls_list(self, raw):
        # Acepta JSON ["u1","u2"] o CSV "u1,u2"
        if raw is None:
            return []
        if isinstance(raw, (list, tuple)):
            return [str(x) for x in raw if x]
        s = str(raw).strip()
        if not s:
            return []
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                return [str(x) for x in parsed if x]
        except Exception:
            pass
        return [p.strip() for p in s.split(',') if p.strip()]

    def _save_extra_images(self, files):
        urls = []
        for f in files or []:
            ruta = self.handle_uploaded_file(f)
            urls.append(ruta)
        return urls

    def _coerce_imagenes_extra_text(self, data_dict):
        # Mantiene compatibilidad con tu helper anterior (si te llega como texto)
        val = data_dict.get('imagenes_adicionales')
        if val is None:
            return []
        if isinstance(val, str):
            s = val.strip()
            try:
                return json.loads(s) if s else []
            except Exception:
                return [p.strip() for p in s.split(',') if p.strip()]
        if isinstance(val, (list, tuple)):
            return [str(x) for x in val if x]
        return []

    def create(self, request, *args, **kwargs):
        """Crear producto (solo admin), incluyendo imágenes adicionales"""
        imagen = request.FILES.get('imagen')
        extra_files = request.FILES.getlist('imagenes_adicionales')  # múltiples
        # Construir dict plano sin archivos
        data = {k: v for k, v in request.data.items() if k != 'imagen'}

        # Lista explícita opcional (reemplazo/orden final inicial)
        explicit_list_raw = request.data.get('imagenes_adicionales_list', None)
        explicit_list = self._parse_urls_list(explicit_list_raw) if explicit_list_raw is not None else []

        # Extra desde texto opcional (JSON/CSV), ya sea en imagenes_adicionales o imagenes_adicionales_urls
        extra_from_text = self._coerce_imagenes_extra_text(data)
        extra_from_text_urls = self._parse_urls_list(request.data.get('imagenes_adicionales_urls'))
        # Extra desde archivos subidos
        extra_from_uploads = self._save_extra_images(extra_files)

        # Construir orden: explícita (si viene) + textos/urls + uploads
        base = explicit_list
        extras = [*base, *extra_from_text, *extra_from_text_urls, *extra_from_uploads]
        if extras or explicit_list_raw is not None:
            # Quita duplicados manteniendo orden (permite lista vacía explícita)
            data['imagenes_adicionales'] = list(dict.fromkeys(x for x in extras if x))

        if imagen:
            ruta = self.handle_uploaded_file(imagen)
            data['imagen_url'] = ruta

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        """Actualizar producto (solo admin), permitiendo reemplazo/orden explícito y agregados"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        imagen = request.FILES.get('imagen')
        extra_files = request.FILES.getlist('imagenes_adicionales')
        data = {k: v for k, v in request.data.items() if k != 'imagen'}

        # Lista explícita opcional (si viene, reemplaza y define orden base)
        explicit_list_raw = request.data.get('imagenes_adicionales_list', None)
        if explicit_list_raw is not None:
            base_list = self._parse_urls_list(explicit_list_raw)
        else:
            base_list = (instance.imagenes_adicionales or [])

        # Permitir compatibilidad: también aceptar texto/urls adicionales
        extra_from_text = self._coerce_imagenes_extra_text(data)
        extra_from_text_urls = self._parse_urls_list(request.data.get('imagenes_adicionales_urls'))
        extra_from_uploads = self._save_extra_images(extra_files)

        combined = [*base_list, *extra_from_text, *extra_from_text_urls, *extra_from_uploads]
        # Si llega lista explícita (aunque sea vacía) o hay algo combinado, seteamos el campo
        if combined or explicit_list_raw is not None:
            data['imagenes_adicionales'] = list(dict.fromkeys(x for x in combined if x))

        if imagen:
            ruta = self.handle_uploaded_file(imagen)
            data['imagen_url'] = ruta

        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

    def handle_uploaded_file(self, f):
        # Guarda el archivo en MEDIA_ROOT/productos/ y retorna la URL relativa bajo MEDIA_URL
        folder = os.path.join(settings.MEDIA_ROOT, 'productos')
        os.makedirs(folder, exist_ok=True)
        base, ext = os.path.splitext(f.name)
        filename = f.name
        path = os.path.join(folder, filename)
        i = 1
        while os.path.exists(path):
            filename = f"{base}_{i}{ext}"
            path = os.path.join(folder, filename)
            i += 1
        with open(path, 'wb+') as destination:
            for chunk in f.chunks():
                destination.write(chunk)
        return os.path.join('media', 'productos', filename).replace('\\', '/')

    # === Exportación a Excel (mover aquí, eliminar el segundo ViewSet) ===
    def _resolve_image_url(self, path, request):
        if not path:
            return ''
        s = str(path)
        if s.lower().startswith('http'):
            return s
        return request.build_absolute_uri('/' + s.lstrip('/'))

    def _safe_text(self, v, max_len=32767):
        if v is None:
            return ''
        if isinstance(v, bool):
            return 'Sí' if v else 'No'
        s = str(v)
        s = ILLEGAL_CHARACTERS_RE.sub('', s)
        return s[:max_len]

    def _safe_num(self, v):
        if v is None or v == '':
            return None
        if isinstance(v, Decimal):
            return float(v)
        try:
            return float(v)
        except Exception:
            return None

    def _safe_dt(self, dt):
        if not dt:
            return None
        # Excel no acepta tz-aware; quita tz
        if getattr(dt, 'tzinfo', None) is not None:
            dt = dt.replace(tzinfo=None)
        return dt

    @action(detail=False, methods=['get'], permission_classes=[IsAdminUser], url_path='export')
    def export(self, request):
        qs = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = 'Productos'

        headers = [
            'ID','Nombre','Categoría','Descripción','Precio','Precio oferta',
            'Stock','Activo','Destacado','Imagen URL',
            'Fecha creación','Última actualización'
        ]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)

        for p in qs:
            categoria = getattr(getattr(p, 'categoria', None), 'nombre', '') or ''
            imagen_path = getattr(p, 'imagen_url', '') or getattr(getattr(p, 'imagen', None), 'url', '')
            img_url = self._resolve_image_url(imagen_path, request)

            row = [
                getattr(p, 'producto_id', getattr(p, 'id', '')),
                self._safe_text(getattr(p, 'nombre', '')),
                self._safe_text(categoria),
                self._safe_text(getattr(p, 'descripcion', '')),
                self._safe_num(getattr(p, 'precio', None)),
                self._safe_num(getattr(p, 'precio_oferta', None)),
                self._safe_num(getattr(p, 'stock', None)),
                self._safe_text(getattr(p, 'activo', False)),
                self._safe_text(getattr(p, 'destacado', False)),
                self._safe_text(img_url),
                self._safe_dt(getattr(p, 'fecha_creacion', None)),
                self._safe_dt(getattr(p, 'fecha_actualizacion', None)),
            ]
            ws.append(row)

        # formatos de fecha (L y M)
        for cell in ws['L'][1:]:
            cell.number_format = 'yyyy-mm-dd hh:mm'
        for cell in ws['M'][1:]:
            cell.number_format = 'yyyy-mm-dd hh:mm'

        # ancho columnas
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(10, min(60, int(max_len * 0.9)))

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"productos_{timezone.now().date().isoformat()}.xlsx"
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp