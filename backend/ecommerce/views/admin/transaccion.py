from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from rest_framework.permissions import IsAdminUser
from ecommerce.models import Transaccion
from ecommerce.serializers import TransaccionSerializer
from django.db import transaction
from rest_framework.response import Response
from rest_framework import status
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from django.core.exceptions import FieldError
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

class TransaccionViewSetAdmin(viewsets.ModelViewSet):
    queryset = Transaccion.objects.select_related('pedido', 'pedido__usuario')
    serializer_class = TransaccionSerializer

    def get_permissions(self):
        return [permissions.IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        base = Transaccion.objects.select_related('pedido', 'pedido__usuario')
        if user.is_staff:
            return base
        return base.filter(pedido__usuario_id=getattr(user, 'usuario_id', None))

    def update(self, request, *args, **kwargs):
        """Al actualizar una transacción, sincroniza el estado del pedido."""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            self.perform_update(serializer)
            # Mapear estado de transacción a estado de pedido
            tx = serializer.instance
            pedido = tx.pedido
            estado_tx = (tx.estado or '').lower()
            if estado_tx in ('aprobado', 'aprobada', 'approved'):
                pedido.estado = 'pagado'
            elif estado_tx in ('rechazado', 'rechazada', 'rejected'):
                pedido.estado = 'rechazado'
            elif estado_tx in ('pendiente', 'pending'):
                # no cambiar
                pass
            elif estado_tx in ('completado', 'completed'):
                pedido.estado = 'completado'
            pedido.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

    # Helpers (si ya los tienes en la clase, puedes omitir estos)
    def _safe_text(self, v, max_len=32767):
        if v is None: return ''
        return ILLEGAL_CHARACTERS_RE.sub('', str(v))[:max_len]

    def _safe_num(self, v):
        if v in (None, ''): return None
        try: return float(v)
        except Exception: return None

    def _safe_dt(self, dt):
        if not dt: return None
        if getattr(dt, 'tzinfo', None) is not None:
            dt = dt.replace(tzinfo=None)
        return dt

    def _parse_date(self, s):
        if not s: return None
        try: return datetime.fromisoformat(str(s)).date()
        except Exception: return None

    def _format_user(self, user, fallback_id=None):
        if not user:
            return f"ID {fallback_id}" if fallback_id else ''
        name = ''
        gf = getattr(user, 'get_full_name', None)
        if callable(gf):
            name = (gf() or '').strip()
        if not name:
            first = getattr(user, 'first_name', '') or getattr(user, 'nombre', '')
            last = getattr(user, 'last_name', '') or getattr(user, 'apellido', '')
            name = f"{first} {last}".strip()
        email = getattr(user, 'email', '') or getattr(user, 'correo', '')
        username = getattr(user, 'username', '') or getattr(user, 'nombre_usuario', '')
        base = name or username
        return self._safe_text(' - '.join([p for p in [base, email] if p]) or (f"ID {fallback_id}" if fallback_id else ''))

    def _resolve_usuario(self, tx):
        user = getattr(tx, 'usuario', None) or getattr(tx, 'user', None) or getattr(tx, 'cliente', None)
        if not user:
            pedido = getattr(tx, 'pedido', None)
            if pedido:
                user = getattr(pedido, 'usuario', None) or getattr(pedido, 'user', None) or getattr(pedido, 'cliente', None)
        fallback_id = (
            getattr(tx, 'usuario_id', None) or getattr(tx, 'user_id', None) or getattr(tx, 'cliente_id', None) or
            (getattr(tx, 'pedido', None) and (
                getattr(tx.pedido, 'usuario_id', None) or getattr(tx.pedido, 'user_id', None) or getattr(tx.pedido, 'cliente_id', None)
            )) or None
        )
        return self._format_user(user, fallback_id)

    @action(detail=False, methods=['get'], permission_classes=[IsAdminUser], url_path='export')
    def export(self, request):
        qs = self.filter_queryset(self.get_queryset())

        # Filtros opcionales: ?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD
        start = self._parse_date(request.query_params.get('start_date'))
        end = self._parse_date(request.query_params.get('end_date'))
        if start:
            qs = qs.filter(fecha_transaccion__date__gte=start)
        if end:
            qs = qs.filter(fecha_transaccion__date__lte=end)

        # Orden por ID asc
        try:
            qs = qs.order_by('transaccion_id')
        except FieldError:
            try:
                qs = qs.order_by('id')
            except FieldError:
                pass

        # Optimiza joins seguros
        try: qs = qs.select_related('pedido')
        except FieldError: pass
        for path in ('pedido__usuario', 'pedido__user', 'pedido__cliente'):
            try:
                qs = qs.select_related(path)
                break
            except FieldError:
                continue

        wb = Workbook()
        ws = wb.active
        ws.title = 'Transacciones'
        headers = ['ID', 'Pedido ID', 'Usuario', 'Monto', 'Método', 'Estado', 'Fecha']
        ws.append(headers)
        for c in ws[1]: c.font = Font(bold=True)

        for t in qs:
            usuario_txt = self._resolve_usuario(t)
            row = [
                getattr(t, 'transaccion_id', getattr(t, 'id', '')),
                getattr(t, 'pedido_id', None) or getattr(getattr(t, 'pedido', None), 'id', ''),
                usuario_txt,
                self._safe_num(getattr(t, 'monto', None)),
                self._safe_text(getattr(t, 'metodo_pago', '')),
                self._safe_text(getattr(t, 'estado', '')),
                self._safe_dt(getattr(t, 'fecha_transaccion', None)),
            ]
            ws.append(row)

        # formato fecha (columna G)
        for cell in ws['G'][1:]:
            cell.number_format = 'yyyy-mm-dd hh:mm'

        # anchos
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(10, min(60, int(max_len * 0.9)))

        bio = BytesIO()
        wb.save(bio); bio.seek(0)

        suffix = []
        if start: suffix.append(f"from_{start.isoformat()}")
        if end:   suffix.append(f"to_{end.isoformat()}")
        filename = "transacciones_" + ("_".join(suffix) if suffix else timezone.now().date().isoformat()) + ".xlsx"

        resp = HttpResponse(
            bio.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp