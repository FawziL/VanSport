from rest_framework import viewsets, permissions
from ecommerce.models import Pedido
from ecommerce.serializers import PedidoSerializer
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from django.core.exceptions import FieldError
from rest_framework.decorators import action
from rest_framework.permissions import IsAdminUser
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

class PedidoViewSetAdmin(viewsets.ModelViewSet):
    queryset = Pedido.objects.select_related('usuario')
    serializer_class = PedidoSerializer

    def get_permissions(self):
        # Solo admin puede ver todos, usuarios normales solo sus pedidos
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        base = Pedido.objects.select_related('usuario')
        if user.is_staff:
            return base
        return base.filter(usuario_id=getattr(user, 'usuario_id', None))

    # Helpers seguros (si ya los tienes en otra clase, puedes reutilizarlos)
    def _safe_text(self, v, max_len=32767):
        if v is None:
            return ''
        return ILLEGAL_CHARACTERS_RE.sub('', str(v))[:max_len]

    def _safe_num(self, v):
        if v in (None, ''):
            return None
        try:
            return float(v)
        except Exception:
            return None

    def _safe_dt(self, dt):
        if not dt:
            return None
        if getattr(dt, 'tzinfo', None) is not None:
            dt = dt.replace(tzinfo=None)
        return dt

    def _parse_date(self, s):
        if not s:
            return None
        try:
            return datetime.fromisoformat(str(s)).date()
        except Exception:
            return None

    def _format_user(self, user, fallback_id=None):
        if not user:
            return f"ID {fallback_id}" if fallback_id else ''
        name = ''
        gf = getattr(user, 'get_full_name', None)
        if callable(gf):
            name = (gf() or '').strip()
        if not name:
            first = getattr(user, 'first_name', '') or getattr(user, 'nombre', '')
            last = getattr(user, 'last_name', '') or getattr(user, 'apellido', '')
            name = f"{first} {last}".strip()
        email = getattr(user, 'email', '') or getattr(user, 'correo', '')
        username = getattr(user, 'username', '') or getattr(user, 'nombre_usuario', '')
        base = name or username
        return self._safe_text(' - '.join([p for p in [base, email] if p]) or (f"ID {fallback_id}" if fallback_id else ''))

    def _resolve_usuario(self, pedido):
        user = getattr(pedido, 'usuario', None) or getattr(pedido, 'user', None) or getattr(pedido, 'cliente', None)
        fallback_id = getattr(pedido, 'usuario_id', None) or getattr(pedido, 'user_id', None) or getattr(pedido, 'cliente_id', None)
        return self._format_user(user, fallback_id)

    @action(detail=False, methods=['get'], permission_classes=[IsAdminUser], url_path='export')
    def export(self, request):
        qs = self.filter_queryset(self.get_queryset())

        # Filtros opcionales: ?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD
        start = self._parse_date(request.query_params.get('start_date'))
        end = self._parse_date(request.query_params.get('end_date'))
        if start:
            qs = qs.filter(fecha_pedido__date__gte=start)
        if end:
            qs = qs.filter(fecha_pedido__date__lte=end)

        # Orden por ID ascendente
        try:
            qs = qs.order_by('pedido_id')
        except FieldError:
            try:
                qs = qs.order_by('id')
            except FieldError:
                pass

        # Optimización segura
        for path in ('usuario', 'user', 'cliente'):
            try:
                qs = qs.select_related(path)
                break
            except FieldError:
                continue

        wb = Workbook()
        ws = wb.active
        ws.title = 'Pedidos'

        headers = ['ID', 'Usuario', 'Fecha', 'Estado', 'Total']
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)

        for p in qs:
            row = [
                getattr(p, 'pedido_id', getattr(p, 'id', '')),
                self._resolve_usuario(p),
                self._safe_dt(getattr(p, 'fecha_pedido', None)),
                self._safe_text(getattr(p, 'estado', '')),
                self._safe_num(getattr(p, 'total', None)),
            ]
            ws.append(row)

        # formato fecha (columna C)
        for cell in ws['C'][1:]:
            cell.number_format = 'yyyy-mm-dd hh:mm'

        # anchos
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(10, min(60, int(max_len * 0.9)))

        bio = BytesIO()
        wb.save(bio); bio.seek(0)

        suffix = []
        if start: suffix.append(f"from_{start.isoformat()}")
        if end:   suffix.append(f"to_{end.isoformat()}")
        filename = "pedidos_" + ("_".join(suffix) if suffix else timezone.now().date().isoformat()) + ".xlsx"

        resp = HttpResponse(
            bio.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp