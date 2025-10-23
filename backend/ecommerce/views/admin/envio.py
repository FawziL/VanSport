from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from rest_framework.permissions import IsAdminUser
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from django.core.exceptions import FieldError
from io import BytesIO
from datetime import datetime
from ecommerce.models import Envio
from ecommerce.serializers import EnvioSerializer

class EnvioViewSetAdmin(viewsets.ModelViewSet):
    queryset = Envio.objects.all()
    serializer_class = EnvioSerializer
    permission_classes = [permissions.IsAdminUser]

    def _safe_text(self, v, max_len=32767):
        if v is None:
            return ''
        s = ILLEGAL_CHARACTERS_RE.sub('', str(v))
        return s[:max_len]

    def _safe_num(self, v):
        if v in (None, ''):
            return None
        try:
            return float(v)
        except Exception:
            return None

    def _safe_dt(self, dt):
        if not dt:
            return None
        if getattr(dt, 'tzinfo', None) is not None:
            dt = dt.replace(tzinfo=None)
        return dt

    def _get_id(self, obj):
        if obj is None:
            return None
        return getattr(obj, 'pk', getattr(obj, 'id', None))

    def _parse_date(self, s):
        # Formato esperado: YYYY-MM-DD
        if not s:
            return None
        s = str(s).strip()
        try:
            return datetime.fromisoformat(s).date()
        except Exception:
            return None

    def _format_user(self, user, fallback_id=None):
        if not user:
            return f"ID {fallback_id}" if fallback_id else ''
        name = ''
        get_full_name = getattr(user, 'get_full_name', None)
        if callable(get_full_name):
            name = (get_full_name() or '').strip()
        if not name:
            first = getattr(user, 'first_name', '') or getattr(user, 'nombre', '')
            last = getattr(user, 'last_name', '') or getattr(user, 'apellido', '')
            name = f"{first} {last}".strip()
        email = getattr(user, 'email', '') or getattr(user, 'correo', '')
        username = getattr(user, 'username', '') or getattr(user, 'nombre_usuario', '')
        parts = [p for p in [name or username, email] if p]
        if not parts and fallback_id:
            parts = [f"ID {fallback_id}"]
        return self._safe_text(' - '.join(parts))

    def _resolve_usuario(self, envio):
        # usuario directo en Envío
        user = getattr(envio, 'usuario', None) or getattr(envio, 'user', None) or getattr(envio, 'cliente', None)
        # si no, tomar desde Pedido
        if not user:
            pedido = getattr(envio, 'pedido', None)
            if pedido:
                user = getattr(pedido, 'usuario', None) or getattr(pedido, 'user', None) or getattr(pedido, 'cliente', None)
        # fallback id
        fallback_id = (
            getattr(envio, 'usuario_id', None) or getattr(envio, 'user_id', None) or getattr(envio, 'cliente_id', None) or
            (getattr(envio, 'pedido', None) and (
                getattr(envio.pedido, 'usuario_id', None) or getattr(envio.pedido, 'user_id', None) or getattr(envio.pedido, 'cliente_id', None)
            )) or None
        )
        return self._format_user(user, fallback_id)

    @action(detail=False, methods=['get'], permission_classes=[IsAdminUser], url_path='export')
    def export(self, request):
        qs = self.filter_queryset(self.get_queryset())

        start = self._parse_date(request.query_params.get('start_date'))
        end = self._parse_date(request.query_params.get('end_date'))
        if start:
            qs = qs.filter(fecha_envio__date__gte=start)
        if end:
            qs = qs.filter(fecha_envio__date__lte=end)

        # Orden por ID ascendente
        try:
            qs = qs.order_by('envio_id')
        except FieldError:
            try:
                qs = qs.order_by('id')
            except FieldError:
                pass

        # Joins seguros
        qs = qs.select_related('pedido')
        # si Pedido tiene usuario/user/cliente, intenta relacionarlo
        for path in ('pedido__usuario', 'pedido__user', 'pedido__cliente'):
            try:
                qs = qs.select_related(path)
                break
            except FieldError:
                continue

        wb = Workbook()
        ws = wb.active
        ws.title = 'Envios'
        headers = ['ID', 'Pedido ID', 'Usuario', 'Método', 'Estado', 'Costo', 'Fecha envío']
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)

        for e in qs:
            # usuario resuelto sin depender del modelo
            usuario_txt = self._resolve_usuario(e)  # <-- cambio clave

            row = [
                getattr(e, 'envio_id', getattr(e, 'id', '')),
                getattr(e, 'pedido_id', None) or getattr(getattr(e, 'pedido', None), 'id', ''),
                self._safe_text(usuario_txt),
                self._safe_text(getattr(e, 'metodo_envio', '')),
                self._safe_text(getattr(e, 'estado', '')),
                self._safe_num(getattr(e, 'costo_envio', None)),
                self._safe_dt(getattr(e, 'fecha_envio', None)),
            ]
            ws.append(row)

        # formato fecha (columna G)
        for cell in ws['G'][1:]:
            cell.number_format = 'yyyy-mm-dd hh:mm'

        # anchos
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(10, min(60, int(max_len * 0.9)))

        bio = BytesIO()
        wb.save(bio); bio.seek(0)

        suffix = []
        if start: suffix.append(f"from_{start.isoformat()}")
        if end:   suffix.append(f"to_{end.isoformat()}")
        name = "envios_" + ("_".join(suffix) if suffix else timezone.now().date().isoformat()) + ".xlsx"

        resp = HttpResponse(
            bio.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="{name}"'
        return resp